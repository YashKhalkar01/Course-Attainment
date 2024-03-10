from django.shortcuts import render, redirect
from django.http import HttpResponse
from collections import Counter,defaultdict
import openpyxl
import io
from home.models import Course
from home.models import Marks
from django.db import connection, connections
from django.contrib.auth import authenticate,login,logout,get_user_model
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages


from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.xml.functions import Element
from openpyxl.xml.constants import (REL_NS, SHEET_MAIN_NS)

@login_required(login_url='user_login')
def home(request):
    return render(request,'home.html')


@login_required(login_url='user_login')
def setPaper(request):
    if request.method == 'POST':
        
        course_code= request.POST.get('course_code')
        branch = request.POST.get('branch')
        table_name = branch+ course_code

        #*********************Creating course tabel*************************
        # Check if the table exists
        existing_tables = connection.introspection.table_names()
        if 'c'+table_name in existing_tables:
            with connections['default'].cursor() as cursor:
                cursor.execute(f"DROP TABLE c{table_name}")
        
        # Create a dynamic model class based on the DynamicTable model
        class NewCourseOutcome(Course):
            class Meta:
                db_table = 'c'+table_name
        
        # Save the dynamic model class in a module-level variable
        globals()[NewCourseOutcome.__name__] = NewCourseOutcome

        # Create the new table
        with connection.schema_editor() as schema_editor:
            schema_editor.create_model(NewCourseOutcome)
        #***********************END*****************************

        #*********************Creating marks tabel*************************
        existing_tables = connection.introspection.table_names()
        if 'm'+table_name in existing_tables:
            with connections['default'].cursor() as cursor:
                cursor.execute(f"DROP TABLE m{table_name}")
              

        # Create a dynamic model class based on the DynamicTable model
        class NewMarks(Marks):
            class Meta:
                db_table = 'm'+table_name
        
        # Save the dynamic model class in a module-level variable
        globals()[NewCourseOutcome.__name__] = NewMarks
        
        # Create the new table
        with connection.schema_editor() as schema_editor:
            schema_editor.create_model(NewMarks)
        #***********************END*****************************

        return render(request,'setPaperSuccessfully.html')

    return render(request,'setPaper.html')


@login_required(login_url='user_login')
def setCourceOutcome(request):   
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        course_name= request.POST.get('course_name')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        #code for storng course outcome
        Q1= request.POST.get('Q1')
        Q2= request.POST.get('Q2')
        Q3= request.POST.get('Q3')
        Q4= request.POST.get('Q4')
        Q5= request.POST.get('Q5')
        Q6= request.POST.get('Q6')
        Q7= request.POST.get('Q7')
        Q8= request.POST.get('Q8')
        table_name=None

        if branch=="Computer Technology":
            table_name="CM"+course_code
        
        elif branch=="Information Technology":
            table_name="IF"+course_code
        
        elif branch=="Automobile Engineering":
            table_name="AE"+course_code
        
        elif branch=="Civil Engineering":
            table_name="CE"+course_code
        
        elif branch=="Electrical Engineering":
            table_name="EE"+course_code
        
        elif branch=="Mechanical Engineering":
            table_name="ME"+course_code
        
        elif branch=="ENTC":
            table_name="ENTC"+course_code
        
        elif branch=="Polymer Engineering":
            table_name="PE"+course_code
        
        elif branch=="Mechatronic Engineering":
            table_name="MCE"+course_code
        
        elif branch=="DDGM":
            table_name="DDGM"+course_code
        
        elif branch=="Interior Design":
            table_name="ID"+course_code
        else : None

        
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT 1 FROM c{table_name} LIMIT 1;") 
            test = cursor.fetchone()
        if test:
            return render(request,'resetCourse.html',{'branch':branch})
        else:
            # Example data to be inserted into the database
            data = {
                'course_name':course_name,
                'branch':branch,
                'year':year,
                'pt':pt,
                'question1': Q1,
                'question2': Q2,
                'question3': Q3,
                'question4': Q4,
                'question5': Q5,
                'question6': Q6,
                'question7': Q7,
                'question8': Q8,
                }
            #******************Insert data into the  course tabel*************
            def insert_data_into_table(table_name, data):
                with connection.cursor() as cursor:
                    # Construct the INSERT query
                    columns = ', '.join(data.keys())
                    placeholders = ', '.join(['%s'] * len(data))
                    query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                    # Execute the query with the given data
                    cursor.execute(query, list(data.values()))
                # Commit the changes to the database
                connection.commit()
            # Call the function to insert the data into the table
            insert_data_into_table('c'+ table_name, data)
            return render(request,'courseSetSuccessfully.html')
    return render(request,'setCourse.html')
    



@login_required(login_url='user_login')
def displayPaper(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')

        #table name
        table_name=None
        
        if branch=="Computer Technology":
            table_name="CM"+course_code
        
        elif branch=="Information Technology":
            table_name="IF"+course_code
        
        elif branch=="Automobile Engineering":
            table_name="AE"+course_code
        
        elif branch=="Civil Engineering":
            table_name="CE"+course_code
        
        elif branch=="Electrical Engineering":
            table_name="EE"+course_code
        
        elif branch=="Mechanical Engineering":
            table_name="ME"+course_code
        
        elif branch=="ENTC":
            table_name="ENTC"+course_code
        
        elif branch=="Polymer Engineering":
            table_name="PE"+course_code
        
        elif branch=="Mechatronic Engineering":
            table_name="MCE"+course_code
        
        elif branch=="DDGM":
            table_name="DDGM"+course_code
        
        elif branch=="Interior Design":
            table_name="ID"+course_code
        else : None

        # Define the columns to be fetched
        columns = ['question1', 'question2', 'question3', 'question4', 'question5', 'question6', 'question7', 'question8']

        # Construct the SQL query
        query2 = "SELECT {} FROM c{} WHERE branch = %s AND year = %s ".format(", ".join(columns), table_name)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query2, [branch, year])
            courseOutcome = cursor.fetchall()
 
            resultArray = [courseOutcome[0][:3],courseOutcome[0][3:8]]
        
        query3 = "SELECT course_name FROM c{} WHERE branch = %s AND year = %s ".format(table_name)    
        with connection.cursor() as cursor:
            cursor.execute(query3, [branch, year])
            courseName = cursor.fetchone()
            
        
        return render(request,'displayPaperSuccessfully.html',{'resultArray':resultArray, 'pt':pt , 'year':year, 'course_code':table_name, 'courseName':courseName,'branch':branch })
    return render(request,'displayPaper.html')



flag=1
@login_required(login_url='user_login')
def excel(request):
    if request.method == 'POST':
        
        course_code = request.POST['course_code']
        branch = str(request.POST.get('branch'))
        year = request.POST.get('year')
        pt =  request.POST.get('pt')
        academic_year =request.POST['academic_year']
        
        #table name
        table_name=None
        
        if branch=="Computer Technology":
            table_name="CM"+course_code
        
        elif branch=="Information Technology":
            table_name="IF"+course_code
        
        elif branch=="Automobile Engineering":
            table_name="AE"+course_code
        
        elif branch=="Civil Engineering":
            table_name="CE"+course_code
        
        elif branch=="Electrical Engineering":
            table_name="EE"+course_code
        
        elif branch=="Mechanical Engineering":
            table_name="ME"+course_code
        
        elif branch=="ENTC":
            table_name="ENTC"+course_code
        
        elif branch=="Polymer Engineering":
            table_name="PE"+course_code
        
        elif branch=="Mechatronic Engineering":
            table_name="MCE"+course_code
        
        elif branch=="DDGM":
            table_name="DDGM"+course_code
        
        elif branch=="Interior Design":
            table_name="ID"+course_code
        else : None

        # Select the active worksheet
        #********************storing marks***************************** 
        # Define the columns to be fetched
        columns = ['roll_no', 'Question1', 'Question2', 'Question3', 'Question4', 'Question5', 'Question6', 'Question7', 'Question8']

        # Construct the SQL query
        query = "SELECT {} FROM m{} WHERE branch = %s AND year = %s ORDER BY roll_no, year".format(", ".join(columns), table_name)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query, [branch, year])
            marks = cursor.fetchall()

        query2 = "SELECT course_name FROM c{} WHERE branch = %s AND year = %s ".format(table_name)    
        with connection.cursor() as cursor:
            cursor.execute(query2, [branch, year])
            course_name = cursor.fetchone()

        courseCodeAndName =course_name[0] +"("+table_name+")"
        
        # Write the results to Excel cells
        file_path='static\Attainment_excel.xlsx'
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        ws.cell(row=5, column=2).value = f" Diploma Programme in {branch}"
        ws.cell(row=7, column=4).value = courseCodeAndName
        ws.cell(row=6, column=4).value = f"Periodic Test-{pt}"
        ws.cell(row=7, column=11).value = academic_year
        i=1
        for row_num, row_data in enumerate(marks):
            for col_num, cell_value in enumerate(row_data):
                ws.cell(row=row_num+12, column=col_num+2).value = cell_value
                ws.cell(row=row_num+12, column=1).value = i
            i=i+1
        wb.save('static/Attainment_excel.xlsx')
      
        #********************************END*****************************
        #**************************upbadting cource outcome***************

        # Define the columns to be fetched
        columns = ['question1', 'question2', 'question3', 'question4', 'question5', 'question6', 'question7', 'question8']

        # Construct the SQL query
        query2 = "SELECT {} FROM c{} WHERE branch = %s AND year = %s ".format(", ".join(columns), table_name)

        # Execute the query and fetch the results
        with connection.cursor() as cursor:
            cursor.execute(query2, [branch, year])
            courseOutcome = cursor.fetchall()
        
        newCourseOutcome = courseOutcome[0]
        counts = Counter(newCourseOutcome)

        # Write the marks to Excel cells
        for rownum, rowdata in enumerate(courseOutcome):
            for colnum, cellvalue in enumerate(rowdata):
                
                ws.cell(row=rownum+10, column=colnum+3).value = cellvalue
               
        # Print the CO Attainment in CO Attainment Average
        row=93
        col=3
        start_col = chr(67)
        end_col = chr(66)
        for element, count in counts.items():
            end_col =chr(ord(end_col)+count)
            
            ws.cell(row=row, column=col).value = element            
            ws.cell(row=96, column=col).value = f'=AVERAGE({start_col}95:{end_col}95)'
            col= col+count 
            start_col =chr(ord(end_col)+1)   
              
         
        with connection.cursor() as cursor:
            # Define the SQL query to fetch the duplicate CO values and their corresponding question numbers
            sql_query = """
                SELECT
                    co_value,
                    GROUP_CONCAT(question_number) AS question_numbers
                FROM (
                    SELECT 1 AS question_number, question1 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 2 AS question_number, question2 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 3 AS question_number, question3 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 4 AS question_number, question4 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 5 AS question_number, question5 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 6 AS question_number, question6 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 7 AS question_number, question7 AS co_value FROM c{table_name}
                    UNION ALL
                    SELECT 8 AS question_number, question8 AS co_value FROM c{table_name}
                ) AS subquery
                WHERE co_value IN ('CO1', 'CO2', 'CO3', 'CO4', 'CO5', 'CO6', 'CO7')
                GROUP BY co_value
                HAVING COUNT(*) > 0
            """.format(table_name=table_name)

            #Execute the SQL query
            cursor.execute(sql_query)

            # Fetch the results
            rows = cursor.fetchall()

        # Process the results and store question numbers as a dictionary of lists
        result = {}
        for row in rows:
            co_value = row[0]
            question_numbers_str = row[1]
            question_numbers = list(map(int, question_numbers_str.split(',')))

            if co_value not in result:
                result[co_value] = []

            # Convert question numbers to the desired format
            question_letters = ['a', 'b', 'c', 'd', 'e']
            global flag
            
            if flag == 1: 
                col=3
            
            for i, question_number in enumerate(question_numbers):
            
                flag = 0
                question_label = f"{'1' if question_number < 4 else '2' }.{question_letters[question_number-1] if question_number < 4 else question_letters[question_number-4]}"
                result[co_value].append(question_label)
                ws.cell(row=95, column=col).value = ws.cell(row=90, column=3+int(question_number-1)).value
               
                col=col+1
                
            wb.save('static/Attainment_excel.xlsx')
        # Display the results
        col=3
        for co_value, question_labels in result.items():
            for i in range(len(question_labels)):
                
                ws.cell(row=94, column=col).value = question_labels[i]
                col=col+1



        # Save the workbook
        wb.save('static/Attainment_excel.xlsx')
       
        #********************************END*****************************


        return render(request,'download.html')
    return render(request,'excel.html')
    
@login_required(login_url='user_login')
def remove(request):
    if request.method == 'POST':
        wb = openpyxl.load_workbook('static\Attainment_excel.xlsx')
       
        # Select the active worksheet
        ws = wb.active

        ws.cell(row=6, column=4).value = None
        ws.cell(row=5, column=2).value = None
        ws.cell(row=7, column=4).value = None
        ws.cell(row=7, column=11).value = None

        for row in ws.iter_rows(min_row=10, min_col=3, max_row=10, max_col=10):
                for cell in row:
                        cell.value = None
        

        for row in ws.iter_rows(min_row=12, min_col=1, max_row=80, max_col=10):
                for cell in row:
                        cell.value = None
       
    
        for row in ws.iter_rows(min_row=93, min_col=3, max_row=96, max_col=12):
                for cell in row:
                        cell.value = None
        wb.save('static\Attainment_excel.xlsx')

        return redirect('home')
                        
    
@login_required(login_url='user_login')
def updateMarks(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        roll_no= request.POST.get('roll_no')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        #code for storng course outcome
        Q1= request.POST.get('Q1')
        Q2= request.POST.get('Q2')
        Q3= request.POST.get('Q3')
        Q4= request.POST.get('Q4')
        Q5= request.POST.get('Q5')
        Q6= request.POST.get('Q6')
        Q7= request.POST.get('Q7')
        Q8= request.POST.get('Q8')
        table_name=None
        if branch=="Computer Technology":
            table_name="CM"+course_code
        
        elif branch=="Information Technology":
            table_name="IF"+course_code
        
        elif branch=="Automobile Engineering":
            table_name="AE"+course_code
        
        elif branch=="Civil Engineering":
            table_name="CE"+course_code
        
        elif branch=="Electrical Engineering":
            table_name="EE"+course_code
        
        elif branch=="Mechanical Engineering":
            table_name="ME"+course_code
        
        elif branch=="ENTC":
            table_name="ENTC"+course_code
        
        elif branch=="Polymer Engineering":
            table_name="PE"+course_code
        
        elif branch=="Mechatronic Engineering":
            table_name="MCE"+course_code
        
        elif branch=="DDGM":
            table_name="DDGM"+course_code
        
        elif branch=="Interior Design":
            table_name="ID"+course_code
        else : None

         # Construct the SQL query to update the table
        query = f"UPDATE m{table_name} SET pt='{pt}', question1='{Q1}', question2='{Q2}', question3='{Q3}', question4='{Q4}', question5='{Q5}', question6='{Q6}', question7='{Q7}', question8='{Q8}' WHERE roll_no='{roll_no}';"

        # Execute the SQL query
        with connection.cursor() as cursor:
            cursor.execute(query)
 
                      
        return render(request,'updateMarksSuccessfully.html')
    return render(request,'updateMarks.html')


@login_required(login_url='user_login')   
def insertMarks(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        roll_no= request.POST.get('roll_no')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        Q1= int(request.POST.get('Q1'))
        Q2= int(request.POST.get('Q2'))
        Q3= int(request.POST.get('Q3'))
        Q4= int(request.POST.get('Q4'))
        Q5= int(request.POST.get('Q5'))
        Q6= int(request.POST.get('Q6'))
        Q7= int(request.POST.get('Q7'))
        Q8= int(request.POST.get('Q8'))
        total=int(request.POST.get('total'))
        #check mark entered correctly
        questions=[[Q1,Q2,Q3],[Q4,Q5,Q6,Q7,Q8]] 
        def test(arry):
            for i in range(0,len(arry)-1):
                for j in range(i+1,len(arry)):
                    if arry[i]<arry[j]:
                        temp=arry[i]
                        arry[i]=arry[j]
                        arry[j]=temp

            if len(arry)==3:
                return arry[0:2]
            elif len(arry)==5:
                return arry[0:3]

        highestQ1=test(questions[0])
                
        highestQ2=test(questions[1])
        totalMarks=sum(highestQ1) + sum(highestQ2)
        if totalMarks !=total:
            messages.success(request, ("Enter accurate marks") )
            return redirect('student')
        else:
            table_name=None
            if branch=="Computer Technology":
                table_name="CM"+course_code
        
            elif branch=="Information Technology":
                table_name="IF"+course_code
            
            elif branch=="Automobile Engineering":
                table_name="AE"+course_code
            
            elif branch=="Civil Engineering":
                table_name="CE"+course_code
            
            elif branch=="Electrical Engineering":
                table_name="EE"+course_code
            
            elif branch=="Mechanical Engineering":
                table_name="ME"+course_code
            
            elif branch=="ENTC":
                table_name="ENTC"+course_code
            
            elif branch=="Polymer Engineering":
                table_name="PE"+course_code
            
            elif branch=="Mechatronic Engineering":
                table_name="MCE"+course_code
            
            elif branch=="DDGM":
                table_name="DDGM"+course_code
            
            elif branch=="Interior Design":
                table_name="ID"+course_code
            else : None
    
            query = f"INSERT INTO m{table_name} (roll_no, branch, year, pt, question1, question2, question3, question4, question5, question6, question7, question8) VALUES ('{roll_no}', '{branch}', '{year}' , '{pt}' ,'{Q1}' ,'{Q2}' ,'{Q3}' ,'{Q4}' ,'{Q5}' ,'{Q6}' ,'{Q7}' ,'{Q8}') "
    
            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(query)
    
            return render(request,'storeMarksSuccessfully.html')
    return render(request,'insertMarks.html')

UserID = 0
@login_required(login_url='user_login')
def change_password(request):
    global UserID
    UserID = request.user.username
    if request.method == 'POST':
        my_user = User.objects.get(username=UserID)
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if confirm_password!=password :
            messages.success(request, ("Password is not Matched") )
            return redirect('change_password')
        
        else :
            my_user.set_password(password)
            if my_user.save():
                messages.success(request, ("Error occured !, Password is not changed ") )
                return redirect('confirm_password')
            else:
                messages.success(request, ("Successfully Password is Changed") )
                return redirect('user_login')
    return render(request, 'authenticate\change_password.html',{'UserID':UserID})

def user_login(request) :
    if request.method == 'POST':

        UserID= request.POST.get('UserID')
        Password= request.POST.get('password')
        user = authenticate(request, username=UserID, password=Password)

        if user is not None:
            login(request, user)
            messages.success(request, ("Successfully Login") )
            return redirect('home')
        else:
            messages.success(request, (" Enter valid user id or password") )

            return redirect('user_login')

    return render(request, 'authenticate\login.html')


def user_signup(request):
    if request.method == 'POST':

        UserID = request.POST.get('UserID')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email')

        if confirm_password!=password :
            messages.success(request, ("Password is not Matched") )
            return redirect('user_signup')

        elif get_user_model().objects.filter(username=UserID).exists():
            messages.success(request, ("User is Already Exist") )
            return redirect('user_signup')
        else:
            my_user= User.objects.create_user(UserID,email,password)
            my_user.save()
            messages.success(request, ("Successfully User is Created") )
            return redirect('user_login')

    return render(request, 'authenticate\signup.html')


def user_logout(request) :
    logout(request)
    messages.success(request, ("Successfully Logout") )
    return redirect('user_login')



def forgot_password(request):
    if request.method == 'POST':
        email=0
        UserID = request.POST.get('UserID')
      
        if not get_user_model().objects.filter(username=UserID).exists():
            messages.success(request, ("User is not  Exist") )
            return redirect('forgot_password')
        else:           
            my_user = User.objects.get(username=UserID)
            email = my_user.email
            password = my_user.password
            return  render(request,'authenticate\email_sent.html',{'email':email})
    return render(request,'authenticate\Forgot_password.html')

def email_sent(request):
    if request.method == 'POST':
        email = request.POST.get('email')
   
    return None 


def student(request):
    if request.method == 'POST':
        course_code= request.POST.get('course_code')
        roll_no= request.POST.get('roll_no')
        branch = request.POST.get('branch')
        year = request.POST.get('year')
        pt = request.POST.get('pt')
        Q1= int(request.POST.get('Q1'))
        Q2= int(request.POST.get('Q2'))
        Q3= int(request.POST.get('Q3'))
        Q4= int(request.POST.get('Q4'))
        Q5= int(request.POST.get('Q5'))
        Q6= int(request.POST.get('Q6'))
        Q7= int(request.POST.get('Q7'))
        Q8= int(request.POST.get('Q8'))
        total=int(request.POST.get('total'))
        #check mark entered correctly
        questions=[[Q1,Q2,Q3],[Q4,Q5,Q6,Q7,Q8]] 
        def test(arry):
            for i in range(0,len(arry)-1):
                for j in range(i+1,len(arry)):
                    if arry[i]<arry[j]:
                        temp=arry[i]
                        arry[i]=arry[j]
                        arry[j]=temp

            if len(arry)==3:
                return arry[0:2]
            elif len(arry)==5:
                return arry[0:3]

        highestQ1=test(questions[0])
                
        highestQ2=test(questions[1])
        totalMarks=sum(highestQ1) + sum(highestQ2)
        if totalMarks !=total:
            messages.success(request, ("Enter accurate marks") )
            return redirect('student')
        else:
            table_name=None

            if branch=="Computer Technology":
                table_name="CM"+course_code
        
            elif branch=="Information Technology":
                table_name="IF"+course_code

            elif branch=="Automobile Engineering":
                table_name="AE"+course_code

            elif branch=="Civil Engineering":
                table_name="CE"+course_code

            elif branch=="Electrical Engineering":
                table_name="EE"+course_code

            elif branch=="Mechanical Engineering":
                table_name="ME"+course_code

            elif branch=="ENTC":
                table_name="ENTC"+course_code

            elif branch=="Polymer Engineering":
                table_name="PE"+course_code

            elif branch=="Mechatronic Engineering":
                table_name="MCE"+course_code

            elif branch=="DDGM":
                table_name="DDGM"+course_code

            elif branch=="Interior Design":
                table_name="ID"+course_code
            else : None

            query = f"INSERT INTO m{table_name} (roll_no, branch, year, pt, question1, question2, question3, question4, question5, question6, question7, question8) VALUES ('{roll_no}', '{branch}', '{year}' , '{pt}' ,'{Q1}' ,'{Q2}' ,'{Q3}' ,'{Q4}' ,'{Q5}' ,'{Q6}' ,'{Q7}' ,'{Q8}') "

            # Execute the SQL query
            with connection.cursor() as cursor:
                cursor.execute(query)

            return render(request,'studentMarksSuccessfully.html')
    return render(request, 'student.html')



#error handling
def error_404_view(request,exception):
    return render(request,'404.html',status=404)

def error_500_view(request):
    return render(request,'500.html',status=500)

def error_400_view(request,exception):
    return render(request,'400.html',status=400)

def error_403_view(request,exception):
    return render(request,'403.html',status=403)
    
